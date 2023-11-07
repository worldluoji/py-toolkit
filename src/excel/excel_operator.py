import openpyxl  

""" copy sheet partly from src_dir to dst_dir
  :param src_dir: source excel file
  :param src_sheetname: source excel's sheetname
  :param src_copy_row: the source cell's rows
  :param src_copy_column: the source cell's columns
  
  :param dst_dir: dst excel file
  :param dst_sheetname: dest excel's sheetname
  :param dst_start_row: dest start row to input content
  :param dst_start_column: dest start column to input content
  :export_file_name: the file to save the result

  :return: None
"""
def copy_sheet(src_dir, src_sheetname, src_copy_row, src_copy_column, 
        dst_dir, dst_sheetname, dst_start_row, dst_start_column, 
        export_file_name="demo.xlsx"):
    
    nrows, src_row_start_index = get_src_row_info(src_copy_row)
    ncolumns, src_column_start_index = get_src_column_info(src_copy_column)

    wbSrc = openpyxl.load_workbook(filename = src_dir)
    wsSrc = wbSrc[src_sheetname]

    wbDst = openpyxl.load_workbook(filename = dst_dir)
    wsDst = wbDst[dst_sheetname]

    last_row = dst_start_row + nrows - 1
    last_columns =  dst_start_column + ncolumns - 1
    rd,cd = wsDst.max_row, wsDst.max_column
    if rd  < last_row:
        wsDst.insert_rows(rd + 1, last_row - rd) # 在第rd 行之后插入last_row - rd 行
    if cd < last_columns:
        wsDst.insert_cols(cd + 1, last_columns - cd)
    
    i, j = src_row_start_index + 1, src_column_start_index + 1
    for r in range(dst_start_row, last_row + 1):
        for c in range(dst_start_column, last_columns + 1):
            wsDst.cell(row=r, column=c).value = wsSrc.cell(row=i, column=j).value
            j += 1
        i += 1
        j = src_column_start_index + 1

    wbDst.save(export_file_name)

""" read successive cells and combine their contents by seperator
  :param filepath: target excel file
  :param sheetname: excel's sheetname
  :param rows: the target cell's rows
  :param columns: the target cell's columns
  :param seperator: use to connect each content in cells
  :return: the content combined by seperator
"""
def read_successive_cells(filepath, sheetname, rows, columns, seperator = "\r\n"):
    nrows, row_start_index = get_src_row_info(rows)
    ncolumns, column_start_index = get_src_column_info(columns)

    wb = openpyxl.load_workbook(filename = filepath)
    ws = wb[sheetname]
    ret = ''
    
    # print(row_start_index + 1, row_start_index + nrows, column_start_index + 1, column_start_index + ncolumns)
    # min_row 等参数下标都是从1开始，不是从0开始
    for row in ws.iter_rows(min_row = row_start_index + 1, max_row = row_start_index + nrows, min_col = column_start_index + 1, max_col = column_start_index + ncolumns):  
        for cell in row:
            if cell.value is not None:
                ret += cell.value + seperator

    return ret.strip(seperator)


"""
    :param source: the excel file path or workbook
    :param sheetname: excel's sheetname
    :cell_values: the list of the cells to be modified, for example: [(1, 2, 'A'), (3, 4, 'B')]
    :return: the workbook of source if succeed else None
"""
def write_to_cells(source, sheetname, cell_values, save_to=''):

    wb = openpyxl.load_workbook(filename = source) if isinstance(source, str) else source
    ws = wb[sheetname]

    r,c = ws.max_row, ws.max_column
    for v in cell_values:
        row,column = v[0], excel_column_alphabet_to_num(v[1])
        if row > r or column > c:
            raise ValueError("row {0} or column {1} out of bounds".format(row, column))
        # 在Openpyxl中，行和列的编号都是从1开始的，而不是从0开始
        ws.cell(row=row, column=column).value = v[2]
   
    return wb if save_excel(wb, save_to) == True else None

        

""" write content to a single cell
  :param source: the excel file path or workbook
  :param sheetname: excel's sheetname
  :param row: the target cell's row
  :param column: the target cell's column
  :return: the workbook of source if succeed else None
"""
def write_to_single_cell(source, sheetname, row, column, content, save_to=''):
    return write_to_cells(source, sheetname, [(row, column, content)], save_to)


""" insert rows into source excel sheet from start_row to next count row
  :param source: the excel file path or workbook
  :param sheetname: excel's sheetname
  :param row: the target cell's row
  :param column: the target cell's column
  :return: the workbook of source if succeed else None
"""
def insert_rows(source, sheetname, start_row, count, cell_values, save_to=''):
    wb = openpyxl.load_workbook(filename = source) if isinstance(source, str) else source
    ws = wb[sheetname]
    try:
        ws.insert_rows(start_row, count)
    except Exception as e:
        print(e)
        return None
    
    return write_to_cells(wb, sheetname, cell_values, save_to)

'''save workbook wb to the dest_path
    :param wb: workbook to be save
    :param dest_path: where to save the workbook
    :return True if no error happend
'''
def save_excel(wb, dest_path):
    if (dest_path is not None) and len(dest_path) > 0:
        try:
            wb.save(dest_path)
        except Exception as e:
            print(e)
            return False
    return True


""" parse the row info
  :param src_copy_row: which rows to be parsed.
  :return nrows: the num of rows
  :return src_copy_row_start - 1: the start index of row
  for example: input "1:2" would return 2,0
"""
def get_src_row_info(src_copy_row):
    src_copy_row_array = src_copy_row.split(":")
    if len(src_copy_row_array) == 1:
        return 1, int(src_copy_row_array[0]) - 1
    src_copy_row_start = int(src_copy_row_array[0])
    src_copy_row_end = int(src_copy_row_array[1])
    nrows = src_copy_row_end - src_copy_row_start + 1
    return nrows, src_copy_row_start - 1
    


excel_col_alphabet_num_map = {
    'A': 1, 'B': 2, 'C': 3, 'D': 4,
    'E': 5, 'F': 6, 'G': 7, 'H': 8,
    'I': 9, 'J': 10, 'K': 11, 'L': 12,
    'M': 13, 'N': 14, 'O': 15, 'P': 16,
    'Q': 17, 'R': 18, 'S': 19, 'T': 20,
    'U': 21, 'V': 22, 'W': 23, 'X': 24,
    'Y': 25, 'Z': 26,
}

BASE = 26

""" parse the column info
  :param src_copy_column: which columns to be parsed.
  :return nrows: the num of columns
  :return  start: the start index of column
  for example: input "A:C" would return 3,0
"""
def get_src_column_info(src_copy_column):
    total = 0
    start = -1
    for part in src_copy_column.split(","):
        src_copy_columns = part.split(":")
        if len(src_copy_columns) == 1:
            total += 1
        else:
            total += excel_column_alphabet_to_num(src_copy_columns[1]) - excel_column_alphabet_to_num(src_copy_columns[0]) + 1
        if start == -1:
            start = excel_column_alphabet_to_num(src_copy_columns[0]) - 1
    return total, start

def excel_column_alphabet_to_num(s):
    if isinstance(s, int):
        return s
    if not isinstance(s, str):
        raise ValueError("s is not int or str")

    c = 0
    for i in range(0, len(s)):
        c += excel_col_alphabet_num_map[s[i]] * pow(BASE, len(s) - i - 1)
    return c