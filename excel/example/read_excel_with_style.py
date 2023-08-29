import openpyxl  
from openpyxl.reader.excel import load_workbook  
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font  

""" This demo shos how to read xlsx with styles, then save to another file
  :file_path:   where the file is
  :sheet_name:  the sheet of the excel file
  
  notification: openpyxl only supports .xlsx file to read with styles
"""
def read_xlsx_file(file_path, sheet_name):  
    # 读取.xlsx文件  
    wb = openpyxl.load_workbook(filename = file_path)
    sheetnames = wb.sheetnames  
    print(sheetnames)
    ws = wb[sheet_name]  
    # 遍历工作表中的单元格并打印其内容和样式  
    for row in ws.iter_rows():  
        for cell in row:  
            print(cell.value, cell.font, cell.fill, cell.border, cell.alignment, cell.protection)

    # 获取最大行数  
    max_row = ws.max_row  
    
    # 获取最大列数  
    max_column = ws.max_column
    print(max_row, max_column)

    wb.save('output.xlsx')

  
# 测试代码  
read_xlsx_file('example.xlsx', 'test')